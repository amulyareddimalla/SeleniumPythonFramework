import json
import logging

import openpyxl
import requests
from datetime import datetime
from selenium.common.exceptions import (
    NoSuchElementException,
    ElementNotVisibleException,
    TimeoutException,
    StaleElementReferenceException,
    InvalidSelectorException,
    WebDriverException
)
from Pages.AddToCartPage import AddToCartPage
from Pages.LogOutPage import LogOutPage
from Pages.LoginPage import LoginPage
# from APImethods.delMethod import DeleteMethod
# from APImethods.getMethod import GetMethod
# from APImethods.postMethod import PostMethod
# from APImethods.putMethod import PutMethod
from utilities import XLUtils
from utilities.BaseClass import BaseClass
from utilities.Excel_Format import apply_formatting_to_status_column_green, \
    apply_formatting_to_status_column_red
from utilities.measure_performance import measure_performance
from utilities.readProperties import ReadConfig
from utilities.webDriverWait import wait_for_element_to_be_clickable, wait_for_element, \
    wait_for_visibility_of_element_located


class TestcaseOne(BaseClass):
    log = BaseClass.getLogger()

    def test_e2e(self):
        readconfig = ReadConfig()  # create object without any argument
        self.baseURL = readconfig.getApplicationURL()
        self.path=readconfig.getinputPath()
        self.output_path=readconfig.getoutputPath()
        workbook = openpyxl.load_workbook("C:/Users/amuly/PycharmProjects1/SeleniumPythonFramework/TestData/login_data.xlsx")
        # workbook = openpyxl.load_workbook("./TestData/login_data.xlsx")
        worksheet = workbook.active
        self.rows = XLUtils.getRowCount(self.path, 'Sheet1')
        # # # Loop through each row in the worksheet
        for r in range(2,self.rows+1):
            # Get the email and password values from the current row
            self.email_value = XLUtils.readData(self.path, 'Sheet1', r, 2)
            self.password_value = XLUtils.readData(self.path, 'Sheet1', r, 3)
            # call the performance utility function
            perf_metrics = measure_performance(self.driver)
            # access the performance metrics as needed
            initial_cpu_usage = perf_metrics["initial_cpu_usage"]
            initial_memory_usage = perf_metrics["initial_memory_usage"]

            self.driver.get(self.baseURL)
            self.log.info("Opened the URL")
            # measure the initial page load time
            initial_load_time = perf_metrics["initial_load_time"]
            self.log.info(f"Initial Load Time: {initial_load_time} ms")

            # access the performance metrics as needed
            post_initial_cpu_usage = perf_metrics["initial_cpu_usage"]
            post_initial_memory_usage = perf_metrics["initial_memory_usage"]

            loginPage = LoginPage(self.driver)
            addtocartpage = AddToCartPage(self.driver)
            logoutPage = LogOutPage(self.driver)

            try:
                # Click on the 'My Account' link
                my_account_link = loginPage.my_account_login()  # self.driver.find_element(By.XPATH, '//a[@title="My Account"]')
                my_account_link.click()
                self.log.info("Clicked on My Account")
                # Wait for the 'Login' button to become clickable and click it
                # Login_button = WebDriverWait(self.driver, 10).until(
                #     EC.element_to_be_clickable(loginPage.login_user()))

                Login_button=wait_for_element_to_be_clickable(self.driver, loginPage.login_user())
                Login_button.click()
                self.log.info("Clicked on Login")
                # Fill in the Login form
                email_input = loginPage.email_user()  # self.driver.find_element(By.NAME, 'email')
                email_input.send_keys(self.email_value)
                password_input = loginPage.password_user()  # self.driver.find_element(By.ID, 'input-password')
                password_input.send_keys(self.password_value)
                submit_button = loginPage.submit_user()  # self.driver.find_element(By.XPATH, '//input[@type="submit"]')
                submit_button.click()
                self.log.info("Entered Email & Password, Clicked on Submit")
                self.log.info("Logged in Successfully")
                # wait for the Phones & PDAs link to be visible and click it
                # phones_pdas_link = WebDriverWait(self.driver, 10).until(
                #     EC.visibility_of_element_located(addtocartpage.phone_pda))
                phones_pdas_link=wait_for_element(self.driver, addtocartpage.phone_pda)
                phones_pdas_link.click()
                self.log.info("Clicked on Phones & PDA's tab")

                # measure the products page load time
                products_load_time = perf_metrics["products_load_time"]
                self.log.info(f"Products Load Time: {products_load_time} ms")

                # get CPU and memory usage after product load
                post_product_cpu_usage = perf_metrics["post_product_cpu_usage"]
                post_product_memory_usage = perf_metrics["post_product_memory_usage"]

                # wait for the iPhone link to be visible and click it
                # iphone_link = WebDriverWait(self.driver, 10).until(
                #     EC.visibility_of_element_located(addtocartpage.phoneTitle))
                iphone_link=wait_for_visibility_of_element_located(self.driver, addtocartpage.phoneTitle)
                iphone_link.click()
                self.log.info("Clicked on iPhone from displayed results")

                # wait for the Add to Cart button to be visible and click it
                # add_to_cart_button = WebDriverWait(self.driver, 10).until(
                #     EC.visibility_of_element_located(addtocartpage.add_to_cart))=
                add_to_cart_button=wait_for_visibility_of_element_located(self.driver, addtocartpage.add_to_cart)
                add_to_cart_button.click()
                self.log.info("Clicked on Add To Cart")

                # wait for the success message to be visible and validate its text
                # success_message = WebDriverWait(self.driver, 10).until(
                #     EC.visibility_of_element_located(
                #         addtocartpage.success_msg))
                success_message=wait_for_visibility_of_element_located(self.driver, addtocartpage.success_msg)
                assert "Success" in success_message.text.strip()
                success_message_text = success_message.text.strip()
                print(success_message_text)
                self.log.info("Validated the success message after adding to cart")

                # Click on Checkout button
                logoutPage.checkout_user().click()
                self.log.info("Clicked on Checkout")

                # Validate if 'Products marked with *** are not available in the desired quantity or not in stock!' message is displayed
                # message = WebDriverWait(self.driver, 10).until(
                #     EC.visibility_of_element_located(addtocartpage.warning_msg))
                message=wait_for_visibility_of_element_located(self.driver, addtocartpage.warning_msg)
                assert message.is_displayed()
                message_output = message.text.strip()
                print(message_output)
                self.log.info("Validated the error message displayed")
                # worksheet.cell(row=2, column=4).value = "Pass"

                XLUtils.writeData(self.output_path, 'Sheet1', r, 4, "Pass")

                # Click on the 'My Account' link
                my_account_link = loginPage.my_account_login()
                my_account_link.click()
                self.log.info("Clicked on My Account for next loop")

                # Wait for the 'Logout' button to become clickable and click it
                # Logout_button = WebDriverWait(self.driver, 10).until(
                #     EC.element_to_be_clickable(logoutPage.logout_user()))
                Logout_button=wait_for_element_to_be_clickable(self.driver, logoutPage.logout_user())
                Logout_button.click()
                self.log.info("Clicked on Logout")

                self.log.info("Initial CPU usage: {}%".format(initial_cpu_usage))
                self.log.info("Post-initial CPU usage: {}%".format(post_initial_cpu_usage))
                self.log.info("Initial memory usage: {}%".format(initial_memory_usage))
                self.log.info("Post-initial memory usage: {}%".format(post_initial_memory_usage))
                self.log.info("Post-product CPU usage: {}%".format(post_product_cpu_usage))
                self.log.info("Post-product memory usage: {}%".format(post_product_memory_usage))

            except (
                    NoSuchElementException,
                    ElementNotVisibleException,
                    TimeoutException,
                    StaleElementReferenceException,
                    InvalidSelectorException,
                    WebDriverException,
                    AssertionError
            ) as e:
                self.log.info(f"{type(e).__name__}: {str(e)}")
                now = datetime.now()
                timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
                screenshot_file = f"C:/Users/amuly/PycharmProjects1/SeleniumPythonFramework/Screenshots/exception_{type(e).__name__}_{timestamp}.png"
                # screenshot_file = f"./SeleniumPythonFramework/Screenshots/exception_{type(e).__name__}_{timestamp}.png"
                self.driver.save_screenshot(screenshot_file)
                self.log.info(f"{type(e).__name__}: {str(e)}\nScreenshot saved to {screenshot_file}")
                XLUtils.writeData(self.output_path, 'Sheet1', r, 4, "Exception")








