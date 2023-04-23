from openpyxl.styles import Font, PatternFill

def apply_formatting_to_status_column_green(cell):
    greenFill = PatternFill(start_color='006400', end_color='006400',  fill_type='solid')
    font = Font(color="FFFFFF")
    if cell.value == "Pass":
        cell.fill = greenFill
        cell.font = font

def apply_formatting_to_status_column_red(cell):
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    font = Font(color="FFFFFF")
    if cell.value == "Fail":
        cell.fill = redFill
        cell.font = font



