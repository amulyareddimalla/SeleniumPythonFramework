
import openpyxl
import requests
import json

from testcases.api import api_URL
from utilities.readProperties import ReadConfig


class PostMethod:
    def post_data(self):
        # readconfig = ReadConfig()  # create object without any argument
        # self.api_URL = readconfig.getAPIURL()
        workbook = openpyxl.load_workbook(
            "C:/Users/amuly/PycharmProjects1/SeleniumPythonFramework/TestData/api_data.xlsx")
        worksheet = workbook.active
        row_count_max = worksheet.max_row

        for row_num in range(2, row_count_max + 1):

            id_value = worksheet.cell(row=row_num, column=1).value
            name = worksheet.cell(row=row_num, column=2).value
            description = worksheet.cell(row=row_num, column=3).value

            try:
                data = json.dumps({'id': id_value, 'name': name, 'description': description})
                response = requests.put(api_URL, data)
                response.raise_for_status()
                print("Request successful with status code: ", + response.status_code)
                print(data)
            except requests.exceptions.RequestException as e:
                print("Request failed:", e)
            except json.JSONDecodeError as e:
                print("Error decoding JSON response:", e)